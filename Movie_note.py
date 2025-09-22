import streamlit as st
import requests
import os, json
from dotenv import load_dotenv
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XLImage
from io import BytesIO
from openpyxl.styles import Alignment, Font, PatternFill
from googleapiclient.discovery import build
import pickle
from google.auth.transport.requests import Request
from datetime import datetime
from io import BytesIO
import io
from googleapiclient.http import MediaFileUpload, MediaIoBaseUpload, MediaIoBaseDownload
from PIL import Image as PILImage
import base64
from google.oauth2 import service_account
from googleapiclient.discovery import build

SCOPES = ["https://www.googleapis.com/auth/drive"]
# SCOPES = ["https://www.googleapis.com/auth/drive.file"]

# .env から API_KEY を読み込み
load_dotenv()
API_KEY = os.getenv("MOVIE_API_KEY")
BASE_URL = "https://api.themoviedb.org/3"
EXCEL_FILE = "movie_note.xlsx"

# =========================================================
# Google Drive サービス取得
# =========================================================

def get_gdrive_service():

    if os.environ.get("GDRIVE_SERVICE_ACCOUNT_JSON"):  # Render等クラウド用
        # 環境変数から読み込む
        service_account_info = json.loads(os.environ["GDRIVE_SERVICE_ACCOUNT_JSON"])
        credentials = service_account.Credentials.from_service_account_info(
            service_account_info, scopes=SCOPES
        )
    else:  # ローカル用
        SERVICE_ACCOUNT_FILE = "service_account.json"
        credentials = service_account.Credentials.from_service_account_file(
            SERVICE_ACCOUNT_FILE, scopes=SCOPES
        )

    return build("drive", "v3", credentials=credentials)

# =========================================================
# Driveからファイルをダウンロード
# =========================================================

def download_from_drive(folder_id, filename="movie_note.xlsx"):
    service = get_gdrive_service()

    # Drive上にファイルがあるか検索
    query = f"'{folder_id}' in parents and name='{filename}' and trashed=false"
    results = service.files().list(q=query, fields="files(id)").execute()
    items = results.get("files", [])

    if not items:
        return None  # ファイルがまだ存在しない

    file_id = items[0]["id"]
    request = service.files().get_media(fileId=file_id)
    fh = io.BytesIO()
    downloader = MediaIoBaseDownload(fh, request)
    done = False
    while not done:
        status, done = downloader.next_chunk()
    fh.seek(0)    
    return fh.getvalue()

# =========================================================
# 見栄えを整える（列幅・行高さ・セル配置など）
# =========================================================

def format_excel(ws):

    # 列幅設定
    col_widths = {
        "A": 20, "B": 20, "C": 10, "D": 15, "E": 20,
        "F": 40, "G": 40, "H": 40
    }
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width

    # 行の高さ：2行目以降はすべて120
    for row in range(2, ws.max_row + 1):
        ws.row_dimensions[row].height = 120

    # A〜H列：縦位置 上詰め
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=8):
        for cell in row:
            cell.alignment = Alignment(vertical="top")

    # E, F列：折り返して表示
    for col in ["E", "F"]:
        for row in range(2, ws.max_row + 1):
            ws[f"{col}{row}"].alignment = Alignment(vertical="top", wrap_text=True)

    # --- ヘッダー行の装飾（1行目） ---
    header_fill = PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid")  # スカイブルー
    header_font = Font(bold=True)

    for cell in ws[1]:  # 1行目の全セル
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = header_font
        cell.fill = header_fill

    return ws

def search_movies(query):
    url = f"{BASE_URL}/search/movie"
    params = {"api_key": API_KEY, "query": query, "language": "ja-JP"}
    res = requests.get(url, params=params)
    return res.json().get("results", [])

def get_movie_details(movie_id, api_key):
    url = f"https://api.themoviedb.org/3/movie/{movie_id}?api_key={api_key}&language=ja-JP&append_to_response=credits"
    response = requests.get(url)
    data = response.json()

    title = data.get("title", "")
    year = data.get("release_date", "")[:4]
    overview = data.get("overview", "")
    director = ""
    if "credits" in data:
        crew = data["credits"].get("crew", [])
        for person in crew:
            if person.get("job") == "Director":
                director = person.get("name", "")
                break

    cast = []
    if "credits" in data:
        cast = [c.get("name", "") for c in data["credits"].get("cast", [])[:3]]

    # ポスターURL
    poster_path = data.get("poster_path")
    poster_url = f"https://image.tmdb.org/t/p/w200{poster_path}" if poster_path else None

    return {
        "タイトル": title,
        "公開年": year,
        "監督": director,
        "出演者": ", ".join(cast),
        "概要": overview,
        "感想": "",  # 入力時に追加
        "ポスター": poster_url
    }

# =========================================================
# EXCELファイルを作成する
# =========================================================

def save_to_excel(movies, folder_id, existing_bytes=None):
    """映画データをExcelに保存し、Google Driveにもアップロードする"""

    if existing_bytes:  # Driveから既存のExcelを取得済みならそれを開く
        wb = load_workbook(filename=BytesIO(existing_bytes))
        ws = wb.active
    elif os.path.exists(EXCEL_FILE):  # ローカルに残っていれば使う
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
    else:  # 完全に新規
        wb = Workbook()
        ws = wb.active
        ws.append(["登録日", "タイトル", "公開年", "監督", "出演者", "概要", "感想", "ポスター"])

    # 画像バイト列を保持しておくリスト（openpyxl が保存時に参照するので生存させる）
    image_streams = []

    today = datetime.now().strftime("%Y-%m-%d")

    for movie in movies:
        # 1行追加（ポスターは空セルにしておく）
        ws.append([
            today,
            movie.get("タイトル", ""),
            movie.get("公開年", ""),
            movie.get("監督", ""),
            movie.get("出演者", ""),
            movie.get("概要", ""),
            movie.get("感想", ""),
            ""  # ポスター列は画像で埋める（H列）
        ])

        # 今追加した行番号
        row_num = ws.max_row

        # ポスター処理はここ（ループ内）
        poster_url = movie.get("ポスター")
        print(f"[DEBUG] row {row_num} poster_url: {poster_url}")

        if poster_url:
            try:
                # ダウンロード（stream=True は任意）
                resp = requests.get(poster_url, timeout=10)
                resp.raise_for_status()

                # BytesIO に読み込み -> PIL でリサイズ -> 再度 BytesIO に保存
                img_data = BytesIO(resp.content)
                pil_img = PILImage.open(img_data)

                # サイズ調整（幅 80 px 例）
                max_width = 80
                if pil_img.width > max_width:
                    ratio = max_width / pil_img.width
                    new_size = (max_width, int(pil_img.height * ratio))
                    pil_img = pil_img.resize(new_size)
                # else: 小さい画像はそのまま

                img_bytes = BytesIO()
                pil_img.save(img_bytes, format="PNG")
                img_bytes.seek(0)

                # 参照を保持しておく（これをしないと保存時に閉じられることがある）
                image_streams.append(img_bytes)

                # openpyxl Image を作ってワークシートに追加
                xl_img = XLImage(img_bytes)
                ws.add_image(xl_img, f"H{row_num}")
                print(f"[DEBUG] ポスター貼付成功: H{row_num}")
            except Exception as e:
                print("[WARN] ポスター画像の取得/処理に失敗:", e)

    # 見栄え整形（必要に応じて format_excel を呼ぶ / ここはあなたの format_excel を使う）
    try:
        format_excel(ws)
    except Exception as e:
        print("[WARN] format_excel でエラー:", e)

    # ローカルに保存（バックアップとして保持）
    wb.save(EXCEL_FILE)

    # --- Google Drive にアップロード ---
    service = get_gdrive_service()

    query = f"'{folder_id}' in parents and name='movie_note.xlsx' and trashed=false"
    results = service.files().list(q=query, fields="files(id)").execute()
    items = results.get("files", [])

    # ローカルファイルを開いてアップロード
    with open(EXCEL_FILE, "rb") as f:
        media = MediaIoBaseUpload(f, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

        if items:
            file_id = items[0]["id"]
            service.files().update(fileId=file_id, media_body=media).execute()
            print("✅ Google Drive 上のファイルを更新しました")
        else:
            file_metadata = {"name": "movie_note.xlsx", "parents": [folder_id]}
            service.files().create(body=file_metadata, media_body=media, fields="id").execute()
            print("✅ Google Drive に新規アップロードしました")

st.title("🎬 映画検索アプリ")

query = st.text_input("映画タイトルを入力してください")
if query:
    results = search_movies(query)

    if results:
        titles = [f"{m['title']} ({m.get('release_date','')[:4]})" for m in results]
        choice = st.radio("検索結果から選択してください:", titles)

        selected = results[titles.index(choice)]
        details = get_movie_details(selected["id"],API_KEY)

        # ポスターを表示（あれば）
        if details.get("ポスター"):
            st.image(details["ポスター"])

        # 監督
        st.write("監督:", details.get("監督", "不明"))

        # 俳優（上位3人）
        st.write("出演者:", details.get("出演者", "不明"))

        # 感想入力エリア
        comment = st.text_area("感想を入力してください")

        # ✅ Streamlit Google Drive保存ボタン 
        if st.button("📤 Google Driveに保存（上書き）"):

            folder_id = "1UNBH5iMlZGyWYEXGZZOfqog2DqS1MkpQ"

            # 1. Driveから既存Excelをダウンロード
            existing_bytes = download_from_drive(folder_id, "movie_note.xlsx")

            # 既存状況のデバッグ表示
            if existing_bytes:
                wb_tmp = load_workbook(filename=BytesIO(existing_bytes))  # ← file= を使う
                st.info(f"DEBUG: 今の最終行（保存前）: {wb_tmp.active.max_row}")
            else:
                st.info("DEBUG: 既存ファイルなし（新規作成）")

            # 3. Drive へ保存（結果も確認表示）
            movie_data = [{
                "タイトル": details.get("タイトル", ""),
                "公開年": details.get("公開年", ""),
                "監督": details.get("監督", ""),
                "出演者": details.get("出演者", ""),
                "概要": details.get("概要", ""),
                "感想": comment,
                "ポスター": details.get("ポスター", None)
            }]
            
            save_to_excel(movie_data, folder_id, existing_bytes=existing_bytes)
            st.success(f"✅ Google Driveに保存しました！")

    else:
        st.warning("検索結果が見つかりませんでした。")




